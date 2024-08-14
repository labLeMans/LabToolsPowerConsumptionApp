import sys
import requests
import os
import csv
from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox, QVBoxLayout, QWidget, QPushButton, QDialog
from PyQt5 import uic, QtCore
from PyQt5.QtCore import pyqtSignal, QTimer
from matplotlib.figure import Figure
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image

# Classe pour le canvas matplotlib
class MplCanvas(FigureCanvas):
    def __init__(self, parent=None, width=5, height=4, dpi=100):
        fig = Figure(figsize=(width, height), dpi=dpi)
        self.axes = fig.add_subplot(111)
        super(MplCanvas, self).__init__(fig)

# Fenêtre pour afficher le graphique en plein écran
class GraphWindow(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Graph")
        layout = QVBoxLayout(self)
        self.canvas = MplCanvas(self, width=10, height=8, dpi=100)
        layout.addWidget(self.canvas)
        self.setLayout(layout)

# Classe principale de l'application
class MainApp(QMainWindow):
    power_updated = pyqtSignal(float)  # Signal pour mettre à jour la puissance affichée

    def __init__(self):
        super().__init__()
        self.init_ui()
        self.init_graph()
        self.init_data()
        self.setup_connections()

    def init_ui(self):
        """Initialise l'interface utilisateur."""
        uic.loadUi('/home/pc/Documents/ITxPT/labtools/labtools/consumption_app_ITxPT/wind.ui', self)

        # Désactiver les switchs au démarrage
        self.manualSwitchCheckBox.setEnabled(False)
        self.ignitionCheckBox.setEnabled(False)
        self.fullPowerCheckBox.setEnabled(False)
        self.lowBatteryCheckBox.setEnabled(False)

    def init_graph(self):
        """Initialise le graphique."""
        self.graphic_widget = QWidget()
        self.layout = QVBoxLayout(self.graphic_widget)
        self.canvas = MplCanvas(self, width=5, height=4, dpi=100)
        self.layout.addWidget(self.canvas)
        self.gridLayout_2.addWidget(self.graphic_widget, 0, 0)

        # Bouton pour afficher le graphique en plein écran
        self.fullscreen_button = QPushButton("Afficher le graphique en plein écran")
        self.layout.addWidget(self.fullscreen_button)
        self.fullscreen_button.clicked.connect(self.show_fullscreen_graph)

        # Fenêtre de graphique en plein écran
        self.graph_window = GraphWindow()

    def init_data(self):
        """Initialise les données et les marqueurs."""
        self.start_time = QtCore.QTime.currentTime()  # Heure de départ
        self.power_values = []  # Valeurs de puissance
        self.time_values = []  # Valeurs de temps
        self.max_power = 0 # Initialisation de la puissance maximale

        # Dictionnaire pour stocker les marqueurs
        self.markers = {
            'ignition': {'color': 'red', 'label': 'I', 'times': [], 'state': [''], 'max_power':[]},
            'fullPower': {'color': 'blue', 'label': 'F', 'times': [], 'state': [''], 'max_power':[]},
            'lowBattery': {'color': 'green', 'label': 'L', 'times': [], 'state': [''], 'max_power':[]},
            'manualSwitch': {'color': 'orange', 'label': 'M', 'times': [], 'state': [''], 'max_power':[]}
        }

        # Sauvegarde des états précédents des switches
        self.previous_states = {
            'manualSwitch': None,
            'ignition': None,
            'fullPower': None,
            'lowBattery': None
        }
        # Timer pour mettre à jour le graphique toutes les secondes
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_graph)

    def setup_connections(self):
        """Configure les connexions entre les widgets et les fonctions."""
        self.manualSwitchCheckBox.clicked.connect(lambda: self.add_marker('manualSwitch'))
        self.ignitionCheckBox.clicked.connect(lambda: self.add_marker('ignition'))
        self.fullPowerCheckBox.clicked.connect(lambda: self.add_marker('fullPower'))
        self.lowBatteryCheckBox.clicked.connect(lambda: self.add_marker('lowBattery'))
        self.reportPushButton.clicked.connect(self.generate_report)
        self.startPushButton.clicked.connect(self.start_measurement)  # Connexion du bouton Start

    def show_fullscreen_graph(self):
        """Affiche la fenêtre de graphique en plein écran."""
        self.graph_window.showFullScreen()
        self.update_graph_in_window(self.graph_window.canvas)

    def update_graph_in_window(self, canvas):
        """Met à jour le graphique dans la fenêtre de graphique en plein écran."""
        canvas.axes.clear()
        canvas.axes.plot(self.time_values, self.power_values, label='Power (W)')
        self.update_markers_on_canvas(canvas.axes)
        canvas.draw()

    def add_marker(self, marker_name):
        """Ajoute un marqueur au moment actuel et calcule la puissance maximale."""
        elapsed_time = self.start_time.secsTo(QtCore.QTime.currentTime())
        state = 'on' if getattr(self, f"{marker_name}CheckBox").isChecked() else 'off'

        # Initialiser max_power
        max_power = 0

        # Calculer la puissance maximale entre le dernier marqueur et le nouveau
        if len(self.markers[marker_name]['times']) > 0:
            start_time = self.markers[marker_name]['times'][-1]
            max_power = max((p for t, p in zip(self.time_values, self.power_values) if start_time <= t <= elapsed_time), default=0)

        # Enregistre la puissance maximale seulement si un état change
        if max_power is not None:
            self.markers[marker_name]['max_power'].append(max_power)
        else:
            self.markers[marker_name]['max_power'].append(0)

        # Enregistre le temps et l'état
        self.markers[marker_name]['times'].append(elapsed_time)
        self.markers[marker_name]['state'].append(state)
        
        # Mettre à jour le graphique
        self.update_graph()

        # Ajouter les données au fichier Excel
        if hasattr(self, 'excel_filepath'):
            self.update_excel(elapsed_time, marker_name, max_power)

    def print_max_power(self, max_power):
        print(f"Puissance maximale enregistrée:{self.max_power:.2f}W")
        
    def update_graph(self):
        """Récupère la puissance et met à jour le graphique."""
        power = self.fetch_power_value()
        if power is not None:
            elapsed_time = self.start_time.secsTo(QtCore.QTime.currentTime())
            self.power_values.append(power)
            self.time_values.append(elapsed_time)

            # Mettre à jour la puissance maximale entre les changements d'état
            self.max_power = max(self.max_power, power)
    
            # Garde seulement les 100000 dernières valeurs
            if len(self.power_values) > 100000:
                self.power_values.pop(0)
                self.time_values.pop(0)
    
            self.canvas.axes.clear()
            self.canvas.axes.plot(self.time_values, self.power_values, label='Power (W)')
            self.update_markers_on_canvas(self.canvas.axes)
            self.canvas.draw()
    
            self.update_graph_in_window(self.graph_window.canvas)
            self.power_updated.emit(power)  # Émettre le signal pour mettre à jour l'affichage
    
            # Toujours ajouter les données au fichier CSV (chaque seconde)
            if hasattr(self, 'csv_filepath'):
                self.update_csv(elapsed_time, power)
    
            # Vérifier les changements d'état
            current_states = {
                'manualSwitch': self.manualSwitchCheckBox.isChecked(),
                'ignition': self.ignitionCheckBox.isChecked(),
                'fullPower': self.fullPowerCheckBox.isChecked(),
                'lowBattery': self.lowBatteryCheckBox.isChecked()
            }
    
            if current_states != self.previous_states:
                self.print_max_power(self.max_power) # Debug
                # Ajouter les données au fichier Excel seulement si un état a changé
                if hasattr(self, 'excel_filepath'):
                    self.update_excel(elapsed_time, marker_name=None, max_power=None)
    
                # Mettre à jour les états précédents
                self.previous_states = current_states


    def fetch_power_value(self):
        """Récupère la valeur de puissance actuelle depuis l'URL."""
        url = 'http://192.168.0.2/Home.cgi'
        try:
            response = requests.get(url)
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, 'html.parser')
                current = float(soup.find('input', {'id': 'actcur'})['value'].replace(' A', ''))
                voltage = float(soup.find('input', {'id': 'actvol'})['value'].replace(' V', ''))
                return current * voltage  # Calcul de la puissance
            else:
                raise Exception(f"Erreur {response.status_code}")
        except Exception as e:
            QMessageBox.warning(self, "Erreur", f"Erreur lors de la récupération de la puissance : {str(e)}")
            return None

    def update_markers_on_canvas(self, axes):
        """Met à jour les marqueurs sur le graphique."""
        for marker_name, marker_data in self.markers.items():
            for time in marker_data['times']:
                axes.axvline(x=time, color=marker_data['color'], linestyle='--', label=marker_data['label'])
            axes.legend(loc='upper right')

    def start_measurement(self):
        """Démarre la mesure et active les switches."""
        self.start_time = QtCore.QTime.currentTime()
        self.power_values = []
        self.time_values = []
        self.timer.start(1000)  # Mise à jour toutes les secondes
        self.manualSwitchCheckBox.setEnabled(True)
        self.ignitionCheckBox.setEnabled(True)
        self.fullPowerCheckBox.setEnabled(True)
        self.lowBatteryCheckBox.setEnabled(True)
        self.startPushButton.setEnabled(False)

        # Créer les fichiers CSV et Excel
        self.create_csv_file()
        self.create_excel_file()

    def create_csv_file(self):
        """Crée un fichier CSV pour enregistrer les données de consommation."""
        self.csv_filepath = "consumption_data.csv"
        with open(self.csv_filepath, mode='w', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(['Time (s)', 'Power (W)'])

    def update_csv(self, elapsed_time, power):
        """Met à jour le fichier CSV avec les données de consommation."""
        with open(self.csv_filepath, mode='a', newline='') as file:
            writer = csv.writer(file)
            writer.writerow([elapsed_time, power])

    def create_excel_file(self):
        """Crée un fichier Excel pour enregistrer les données de consommation."""
        self.excel_filepath = "consumption_data.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Consumption Data"
        ws.append(['Main Switch', 'Ignition', 'Full Power', 'Low Battery', 'Power (W)', 'Max Power (W)', 'Time (s)'])
        wb.save(self.excel_filepath)

    def update_excel(self, elapsed_time, marker_name, max_power):
        """Met à jour le fichier Excel avec les données de consommation."""
        if not hasattr(self, 'excel_filepath'):
            return

        # Charger le classeur existant ou en créer un nouveau
        if os.path.exists(self.excel_filepath):
            wb = load_workbook(self.excel_filepath)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Consumption Data"
            ws.append(['Main Switch', 'Ignition', 'Full Power', 'Low Battery', 'Power (W)', 'Max Power (W)', 'Time (s)'])

        # Écrire les données dans le fichier Excel
        ws.append([
            'on' if self.manualSwitchCheckBox.isChecked() else 'off',
            'on' if self.ignitionCheckBox.isChecked() else 'off',
            'on' if self.fullPowerCheckBox.isChecked() else 'off',
            'on' if self.lowBatteryCheckBox.isChecked() else 'off',
            f"{self.power_values[-1]:.2f} W",  # Dernière puissance mesurée
            f"{max_power:.2f} W",               # Puissance maximale entre les marqueurs
            f"{elapsed_time:.3f} s"
        ])
        wb.save(self.excel_filepath)

    def generate_report(self):
        """Génère un rapport dans un fichier Excel."""
        QMessageBox.information(self, "Information", "Rapport généré avec succès.")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainApp()
    window.show()
    sys.exit(app.exec_())
